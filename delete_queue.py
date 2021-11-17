import requests
import json
from openpyxl import load_workbook
from credentials import credentials

list_url = 'https://webexapis.com/v1/telephony/config/queues'
delete_url_portion = 'https://webexapis.com/v1/telephony/config/locations/'
bearer = credentials['bearer']
headers = {'Authorization': 'Bearer ' + bearer}
filename = credentials['filename']

response = requests.get(list_url, headers=headers)
response = response.json()

print(response)

l = len(response['queues'])
print(l)


for i in range(l):
    queue_id = response['queues'][i]['id']
    location_id = response['queues'][i]['locationId']
    print(queue_id)
    print(location_id)
    delete_url = delete_url_portion + location_id+'/queues/' + queue_id
    print(delete_url)
    resp = requests.delete(delete_url, headers=headers)
    print (resp)

#remove licenses
license_url = 'https://webexapis.com/v1/licenses'

response = requests.get(license_url, headers=headers)
response = response.json()
licenses_list = response['items']
l = len(licenses_list)
for i in range (l):
    if licenses_list[i]['name'] == "Webex Calling - Professional":
        wxc_license = licenses_list[i]['id']
        break
people_url = 'https://webexapis.com/v1/people/'
user_details_url = 'https://webexapis.com/v1/people?email='
location_id_url = 'https://webexapis.com/v1/locations'
headers = {'Authorization': 'Bearer ' + bearer}


workbook = load_workbook(filename = filename)
sheet = workbook.active
max_rows_per_column = []
col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
print(col_names)

#calculate the number of items for each column
l = len (col_names)
for i in range (l):
    letter = chr(i+65)
    max_rows = max((c.row for c in sheet[letter] if c.value is not None))
    print ('max rows: ', max_rows)
    max_rows_per_column.append (max_rows)
#create a list where each item is the list of emails for the corresponding column
array=[]
for queue in range(l):
    email_list = []
    for i in range(max_rows_per_column[queue]):
        email_list.append(sheet.cell (i+1, queue+1).value)

    array.append(email_list)
print(array)

queue_number = len(array)
for i in range(queue_number):

    queue_name = array[i][0]
    agent_list = array[i][1:]
    n_agents = len (agent_list)
    agent_id_list = []

    for j in range(n_agents):
       email = agent_list[j]
       #get user ID
       response = requests.get(user_details_url+email, headers=headers)
       response = response.json()
       print(response)
       agent_id = response['items'][0]['id']
       agent_licenses = response['items'][0]['licenses']
       if wxc_license in agent_licenses:
           agent_licenses.remove(wxc_license)
           data = response['items'][0]
           if 'displayName' not in data.keys():
               print("displayName not in response")
               data['displayName'] = ""
           #phoneNumber= response['items'][0]['phoneNumbers'][0]['value']
           #extension = phoneNumber[-4:]
           #del data['phoneNumbers']
           data['licenses'] = agent_licenses
           #data['extension'] = extension
           #data['locationId'] = location_id
           content_type_headers = headers
           content_type_headers['Content-Type'] = 'application/json'
           payload = json.dumps(data)
           print('payload is: ', payload)
           url = people_url + data['id']+ '?callingData=true'
           print('sending request to: ', url)
           response = requests.put(people_url + data['id']+ '?callingData=true', data=payload, headers=content_type_headers)
           print(response)
           print('put response is: ', response.json())
       agent_id_list.append({'id': agent_id})
    print(agent_id_list)
