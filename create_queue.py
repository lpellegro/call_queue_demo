#this script adds users to a queue
import requests
import json
from openpyxl import load_workbook
from credentials import credentials

bearer = credentials['bearer']
filename = credentials['filename']

people_url = 'https://webexapis.com/v1/people/'
user_details_url = 'https://webexapis.com/v1/people?email='
location_id_url = 'https://webexapis.com/v1/locations'
headers = {'Authorization': 'Bearer ' + bearer}
license_url = 'https://webexapis.com/v1/licenses'

response = requests.get(license_url, headers=headers)
response = response.json()
licenses_list = response['items']
l = len(licenses_list)
for i in range (l):
    if licenses_list[i]['name'] == "Webex Calling - Professional":
        wxc_license = licenses_list[i]['id']
        break



workbook = load_workbook(filename = filename)
sheet = workbook.active
max_rows_per_column = []
col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
print(col_names)

#calculate the number of items for each column
l = len (col_names)
for i in range (l):
    letter = chr(i+65)
    max_rows = max((c.row for c in sheet[letter] if c.value is not None))
    print ('max rows: ', max_rows)
    max_rows_per_column.append (max_rows)
#create a list where each item is the list of emails for the corresponding column
array=[]
for queue in range(l):
    email_list = []
    for i in range(max_rows_per_column[queue]):
        email_list.append(sheet.cell (i+1, queue+1).value)

    array.append(email_list)
print(array)

queue_number = len(array)
for i in range(queue_number):

    queue_name = array[i][0]
    agent_list = array[i][1:]
    n_agents = len (agent_list)
    agent_id_list = []
    response = requests.get(location_id_url, headers=headers)
    response = response.json()
    print("get_location_response is:", response)
    location_id = response['items'][0]['id']
    print(location_id)
    for j in range(n_agents):
       email = agent_list[j]
       #get user ID
       response = requests.get(user_details_url+email, headers=headers)
       response = response.json()
       print("user's details are:", response)
       agent_id = response['items'][0]['id']
       agent_licenses = response['items'][0]['licenses']
       if wxc_license not in agent_licenses:
           data = response['items'][0]
           agent_licenses.append(wxc_license)
           if 'phoneNumbers' not in data.keys():
               print("phoneNumbers not in response")
               extension = current_extension
               current_extension += 1
               print("considering extension:", extension)
           else:
               phoneNumber= response['items'][0]['phoneNumbers'][0]['value']
               print("phoneNumbers in response")
               extension = phoneNumber[-4:]
               print("considering extension:", extension)
               del data['phoneNumbers']

           if 'displayName' not in data.keys():
               print("displayName not in response")
               data['displayName'] = ""
           data['licenses'] = agent_licenses
           data['extension'] = extension
           data['locationId'] = location_id
           content_type_headers = headers
           content_type_headers['Content-Type'] = 'application/json'
           payload = json.dumps(data)
           print('payload is: ', payload)
           url = people_url + data['id']+ '?callingData=true'
           print('sending request to: ', url)
           response = requests.put(people_url + data['id']+ '?callingData=true', data=payload, headers=content_type_headers)
           print(response)
           print('put response is: ', response.json())
       agent_id_list.append({'id': agent_id})
    print(agent_id_list)

    #create a queue and add users to the queue

    create_queue_url = 'https://webexapis.com/v1/telephony/config/locations/'+ location_id + '/queues'
    queue_extension = credentials["queue_starting_number"] + i
    print("creating queue with extension:", queue_extension)
    data = json.dumps (
     {
        "name": queue_name,
        "extension": queue_extension,
        "callPolicies": {
            "policy": "UNIFORM",
            "callBounce": {
                "callBounceEnabled": True
            },
            "distinctiveRing": {
                "enabled": False
            }
        },
        "queueSettings": {
            "queueSize": 5,
            "overflow": {
                "action": "PERFORM_BUSY_TREATMENT",
                "greeting": "CUSTOM"
            },



            "enabled": True
            },
        "agents":agent_id_list
     }

    )
    headers['Content-Type'] = 'application/json'

    print (headers)
    #create a queue and assign agents to the queue
    response = requests.post(create_queue_url, headers=headers, data=data)
    print(response.text)
    response.json()
    print(response)

